import os
import glob
import json
import re
import uuid
import shutil
import logging
from datetime import datetime
from typing import Optional

from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.config import Config
from app.database.db_manager import get_db
from app.services.ocr_service import OCRService
from app.services.grading_service import GradingService
from app.services.report_service import ReportService
from app.services.ai_grading_service import AIGradingService
from app.services.detailed_report_service import DetailedReportService
from pypdf import PdfReader

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Smart Grading API", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize services
ocr_service = OCRService()
grading_service = GradingService()
report_service = ReportService(Config.REPORT_DIR)
detailed_report_service = DetailedReportService(Config.REPORT_DIR)

_ai_grading_service = None


def get_ai_grading_service() -> AIGradingService:
    """Lazy-init supaya server tetap bisa jalan meski GEMINI_API_KEY belum diisi,
    error baru muncul saat endpoint AI grading benar-benar dipanggil."""
    global _ai_grading_service
    if _ai_grading_service is None:
        _ai_grading_service = AIGradingService()
    return _ai_grading_service


def _extract_pdf_text(pdf_path: str) -> str:
    """Ekstrak teks dari PDF digital (bukan hasil scan), dipakai untuk rubrik."""
    reader = PdfReader(pdf_path)
    return "\n".join(page.extract_text() or "" for page in reader.pages)


@app.get("/")
def read_root():
    return {"message": "Smart Grading API is running"}


@app.post("/api/upload/answersheet")
async def upload_answer_sheet(
    exam_id: str = Form(...),
    student_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Upload jawaban SATU mahasiswa (dipanggil PHP setelah upload manual lewat
    form 'Upload Jawaban Mahasiswa'). Disimpan ke folder ujian yang sama dengan
    upload massal, dan path resminya disinkronkan ke kolom submissions.answer_sheet_path
    supaya bisa langsung dipakai proses 'Analisis Keseluruhan'."""
    try:
        allowed_ext = ('.pdf', '.jpg', '.jpeg', '.png')
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_ext:
            raise HTTPException(status_code=400, detail="Hanya file PDF, JPG, JPEG, atau PNG yang didukung.")

        exam_row = db.execute(
            text("SELECT id, title, class_id, upload_folder FROM exams WHERE id = :exam_id"),
            {"exam_id": int(exam_id)}
        ).fetchone()
        if not exam_row:
            raise HTTPException(status_code=404, detail="Ujian tidak ditemukan.")

        student_row = db.execute(
            text("""
                SELECT s.name, cs.absen
                FROM students s
                LEFT JOIN class_students cs ON cs.student_id = s.id AND cs.class_id = :class_id
                WHERE s.id = :student_id
            """),
            {"class_id": exam_row.class_id, "student_id": int(student_id)}
        ).fetchone()

        exam_folder = _get_exam_upload_folder(int(exam_id), exam_row.title, exam_row.upload_folder)
        if not exam_row.upload_folder:
            db.execute(
                text("UPDATE exams SET upload_folder = :folder WHERE id = :exam_id"),
                {"folder": os.path.basename(exam_folder), "exam_id": int(exam_id)}
            )

        absen_prefix = student_row.absen if (student_row and student_row.absen is not None) else "X"
        student_name = student_row.name if student_row else "mahasiswa"
        safe_name = _slugify_title(student_name)
        saved_filename = f"{absen_prefix}_{safe_name}{file_ext}"
        file_path = os.path.join(exam_folder, saved_filename)

        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        existing = db.execute(
            text("SELECT id FROM submissions WHERE exam_id = :exam_id AND student_id = :student_id"),
            {"exam_id": int(exam_id), "student_id": int(student_id)}
        ).fetchone()

        if existing:
            db.execute(
                text("UPDATE submissions SET answer_sheet_path = :path WHERE id = :id"),
                {"path": file_path, "id": existing.id}
            )
            db.commit()

        return {
            "status": "success",
            "filename": saved_filename,
            "path": file_path
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error uploading answer sheet: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _parse_rubric_file(rubric_path: str) -> dict:
    """Parse a simple 'key=value' rubric text file. Falls back to sensible
    defaults for anything missing or if the file can't be parsed."""
    rubric = {"max_score": 100.0, "keyword_weight": 0.6, "similarity_weight": 0.4}

    if not rubric_path or not rubric_path.lower().endswith('.txt'):
        return rubric

    try:
        with open(rubric_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                line = line.strip()
                if not line or '=' not in line:
                    continue
                key, value = line.split('=', 1)
                key = key.strip()
                value = value.strip()
                if key in ('max_score', 'keyword_weight', 'similarity_weight'):
                    try:
                        rubric[key] = float(value)
                    except ValueError:
                        pass
    except Exception as e:
        logger.warning(f"Could not parse rubric file, using defaults: {e}")

    return rubric


def _slugify_title(title: str) -> str:
    """Ubah judul ujian jadi nama folder yang aman & mudah dibaca manusia."""
    slug = re.sub(r'[^\w\s-]', '', title, flags=re.UNICODE).strip()
    slug = re.sub(r'[\s]+', '_', slug)
    return slug[:80] if slug else "ujian"


def _get_exam_upload_folder(exam_id: int, title: str, upload_folder_db_value: Optional[str]) -> str:
    """Kembalikan path folder upload untuk suatu ujian. Kalau sudah pernah
    dibuat (tersimpan di kolom exams.upload_folder), pakai itu supaya stabil
    walau judul ujian diedit belakangan. Kalau belum ada, buat baru.
    SELALU absolut, supaya path yang disimpan ke submissions.answer_sheet_path
    juga absolut (Config.UPLOAD_DIR bisa jadi cuma './uploads' yang relatif)."""
    if upload_folder_db_value:
        folder_path = os.path.join(Config.UPLOAD_DIR, "exams", upload_folder_db_value)
    else:
        folder_name = f"{exam_id}_{_slugify_title(title)}"
        folder_path = os.path.join(Config.UPLOAD_DIR, "exams", folder_name)
    folder_path = os.path.abspath(folder_path)
    os.makedirs(folder_path, exist_ok=True)
    return folder_path


def _debug_list_folder(folder_path: str, limit: int = 15) -> str:
    """Bantuan diagnostik: kalau sebuah file tidak ketemu, tunjukkan isi
    folder tempat harusnya file itu berada, supaya gampang ketahuan
    penyebabnya (folder tidak ada, nama beda, dsb)."""
    try:
        if not os.path.isdir(folder_path):
            return f"[Folder '{folder_path}' tidak ada sama sekali di server.]"
        entries = os.listdir(folder_path)
        if not entries:
            return f"[Folder '{folder_path}' ada tapi kosong.]"
        shown = entries[:limit]
        more = f" (+{len(entries) - limit} lainnya)" if len(entries) > limit else ""
        return f"Isi folder '{folder_path}': {', '.join(shown)}{more}"
    except Exception as e:
        return f"[Gagal membaca folder '{folder_path}': {e}]"


def _find_submission_file(exam_id: str, student_id: str) -> Optional[str]:
    """[LEGACY] Cari file jawaban dengan pola folder lama (exam_id/student_id/*.pdf),
    dipakai sebagai fallback untuk submission yang dibuat sebelum skema folder baru."""
    pattern = os.path.join(Config.UPLOAD_DIR, str(exam_id), str(student_id), "*.pdf")
    matches = glob.glob(pattern)
    return matches[0] if matches else None


def _resolve_answer_sheet_path(exam_id: int, student_id: int, answer_sheet_path: Optional[str]) -> Optional[str]:
    """Cari file fisik jawaban mahasiswa. Coba berurutan:
    1. Path absolut yang sudah tersimpan di kolom submissions.answer_sheet_path (skema baru, normal).
    2. Path yang sama tapi ternyata tersimpan RELATIF (bug lama) — coba resolve
       relatif terhadap working directory proses backend saat ini.
    3. Fallback ke pola folder lama (exam_id/student_id/*.pdf) untuk data yang
       dibuat sebelum skema folder-per-ujian ada."""
    if answer_sheet_path:
        if os.path.isabs(answer_sheet_path) and os.path.exists(answer_sheet_path):
            return answer_sheet_path
        candidate = os.path.abspath(answer_sheet_path)
        if os.path.exists(candidate):
            return candidate
    return _find_submission_file(str(exam_id), str(student_id))


@app.post("/api/process/exam")
async def process_exam(
    exam_id: str = Form(...),
    answer_key: UploadFile = File(...),
    rubric: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    """Process grading for every pending/processing submission of an exam:
    OCR the answer key, OCR each student's answer sheet, grade it, and
    persist the score back to the database."""
    try:
        # 1. Simpan kunci jawaban
        exam_upload_dir = os.path.join(Config.UPLOAD_DIR, exam_id)
        os.makedirs(exam_upload_dir, exist_ok=True)

        key_path = os.path.join(exam_upload_dir, "answer_key.pdf")
        with open(key_path, "wb") as buffer:
            shutil.copyfileobj(answer_key.file, buffer)

        # 2. OCR kunci jawaban
        key_text = ocr_service.extract_text_from_pdf(key_path)

        # 3. Simpan & parse rubrik (opsional)
        rubric_data = {"max_score": 100.0, "keyword_weight": 0.6, "similarity_weight": 0.4}
        if rubric is not None and rubric.filename:
            rubric_path = os.path.join(exam_upload_dir, f"rubric_{rubric.filename}")
            with open(rubric_path, "wb") as buffer:
                shutil.copyfileobj(rubric.file, buffer)
            rubric_data = _parse_rubric_file(rubric_path)

        # 4. Ambil submission yang menunggu diproses
        submissions = db.execute(
            text("""
                SELECT id, student_id
                FROM submissions
                WHERE exam_id = :exam_id AND status IN ('pending', 'processing')
            """),
            {"exam_id": exam_id}
        ).fetchall()

        processed = 0
        failed = 0
        results = []

        for sub in submissions:
            submission_id = sub.id
            student_id = sub.student_id

            try:
                answer_sheet_path = _find_submission_file(exam_id, student_id)

                if not answer_sheet_path:
                    raise Exception(
                        "File jawaban tidak ditemukan di backend. "
                        "Pastikan submission diupload lewat frontend terlebih dahulu."
                    )

                # OCR jawaban mahasiswa
                student_text = ocr_service.extract_text_from_pdf(answer_sheet_path)

                # Nilai jawaban
                grading_result = grading_service.grade_answer(student_text, key_text, rubric_data)

                # Simpan nilai ke tabel scores
                db.execute(
                    text("""
                        INSERT INTO scores (submission_id, question_number, score, max_score, feedback)
                        VALUES (:submission_id, 1, :score, :max_score, :feedback)
                    """),
                    {
                        "submission_id": submission_id,
                        "score": grading_result["score"],
                        "max_score": rubric_data["max_score"],
                        "feedback": grading_result["feedback"],
                    }
                )

                # Update status submission
                db.execute(
                    text("""
                        UPDATE submissions
                        SET status = 'completed', processed_at = :processed_at
                        WHERE id = :submission_id
                    """),
                    {"processed_at": datetime.utcnow(), "submission_id": submission_id}
                )

                db.commit()
                processed += 1
                results.append({
                    "submission_id": submission_id,
                    "student_id": student_id,
                    "score": grading_result["score"],
                    "status": "completed"
                })

            except Exception as sub_error:
                db.rollback()
                logger.error(f"Failed grading submission {submission_id}: {sub_error}")

                db.execute(
                    text("UPDATE submissions SET status = 'failed' WHERE id = :submission_id"),
                    {"submission_id": submission_id}
                )
                db.commit()

                failed += 1
                results.append({
                    "submission_id": submission_id,
                    "student_id": student_id,
                    "status": "failed",
                    "error": str(sub_error)
                })

        return {
            "status": "completed",
            "exam_id": exam_id,
            "total_submissions": len(submissions),
            "processed": processed,
            "failed": failed,
            "results": results
        }

    except Exception as e:
        logger.error(f"Error processing exam {exam_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/grade/student")
async def grade_student_answer(
    student_answer: str = Form(...),
    key_answer: str = Form(...),
    max_score: float = Form(100)
):
    """Grade a single student answer (utility endpoint, not tied to DB)."""
    try:
        rubric = {
            "max_score": max_score,
            "keyword_weight": 0.6,
            "similarity_weight": 0.4
        }

        result = grading_service.grade_answer(student_answer, key_answer, rubric)
        return {
            "status": "success",
            "result": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/report/{exam_id}")
async def generate_report(exam_id: str, db: Session = Depends(get_db)):
    """Generate a PDF grading report for an exam, using real data from the database."""
    try:
        exam_row = db.execute(
            text("""
                SELECT e.title, c.name as class_name
                FROM exams e
                LEFT JOIN classes c ON e.class_id = c.id
                WHERE e.id = :exam_id
            """),
            {"exam_id": exam_id}
        ).fetchone()

        if not exam_row:
            raise HTTPException(status_code=404, detail="Exam not found")

        exam_data = {
            "title": exam_row.title,
            "class_name": exam_row.class_name or "N/A"
        }

        score_rows = db.execute(
            text("""
                SELECT s.name as student_name, s.nim,
                       COALESCE(AVG(sc.score), 0) as score
                FROM submissions sub
                JOIN students s ON sub.student_id = s.id
                LEFT JOIN scores sc ON sc.submission_id = sub.id
                WHERE sub.exam_id = :exam_id
                GROUP BY s.name, s.nim
                ORDER BY s.name
            """),
            {"exam_id": exam_id}
        ).fetchall()

        scores = [
            {"student_name": row.student_name, "nim": row.nim, "score": float(row.score)}
            for row in score_rows
        ]

        if not scores:
            raise HTTPException(status_code=400, detail="No submissions found for this exam yet")

        report_path = report_service.generate_grading_report(exam_data, scores)

        return FileResponse(
            report_path,
            media_type='application/pdf',
            filename=os.path.basename(report_path)
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating report for exam {exam_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _grade_submission_core(submission_id: int, db: Session, ai_service) -> dict:
    """Logika inti penilaian AI untuk satu submission. Melempar Exception biasa
    (bukan HTTPException) supaya bisa dipakai baik oleh endpoint single maupun
    endpoint batch ('Analisis Keseluruhan') tanpa saling mengganggu."""
    submission_row = db.execute(
        text("""
            SELECT sub.id, sub.exam_id, sub.student_id, sub.answer_sheet_path,
                   e.title as exam_title, e.rubric_path, e.answer_key_path,
                   s.name as student_name, s.nim as student_nim,
                   c.name as class_name
            FROM submissions sub
            JOIN exams e ON sub.exam_id = e.id
            JOIN students s ON sub.student_id = s.id
            LEFT JOIN classes c ON e.class_id = c.id
            WHERE sub.id = :submission_id
        """),
        {"submission_id": submission_id}
    ).fetchone()

    if not submission_row:
        raise Exception("Submission not found")

    if not submission_row.rubric_path:
        raise Exception("Ujian ini belum punya file rubrik (Markdown). Upload rubrik dulu lewat halaman Edit Ujian.")

    if not submission_row.answer_key_path:
        raise Exception("Ujian ini belum punya file kunci jawaban (Markdown). Upload dulu lewat halaman Edit Ujian.")

    frontend_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "frontend"))
    rubric_full_path = os.path.join(frontend_root, submission_row.rubric_path)
    answer_key_full_path = os.path.join(frontend_root, submission_row.answer_key_path)

    if not os.path.exists(rubric_full_path):
        raise Exception(
            f"File rubrik tidak ditemukan di server: {submission_row.rubric_path} "
            f"(dicari di path lengkap: {rubric_full_path}). {_debug_list_folder(os.path.dirname(rubric_full_path))}"
        )
    if not os.path.exists(answer_key_full_path):
        raise Exception(
            f"File kunci jawaban tidak ditemukan di server: {submission_row.answer_key_path} "
            f"(dicari di path lengkap: {answer_key_full_path}). {_debug_list_folder(os.path.dirname(answer_key_full_path))}"
        )

    answer_sheet_path = _resolve_answer_sheet_path(
        submission_row.exam_id, submission_row.student_id, submission_row.answer_sheet_path
    )
    if not answer_sheet_path:
        raise Exception("File jawaban mahasiswa tidak ditemukan di backend. Upload ulang lewat halaman detail ujian.")

    with open(rubric_full_path, 'r', encoding='utf-8', errors='ignore') as f:
        rubric_text = f.read()
    with open(answer_key_full_path, 'r', encoding='utf-8', errors='ignore') as f:
        answer_key_text = f.read()

    if not rubric_text.strip():
        raise Exception("File rubrik kosong.")
    if not answer_key_text.strip():
        raise Exception("File kunci jawaban kosong.")

    grading_result = ai_service.grade_exam(
        answer_sheet_pdf_path=answer_sheet_path,
        rubric_text=rubric_text,
        answer_key_text=answer_key_text,
        student_name=submission_row.student_name,
        student_nim=submission_row.student_nim
    )

    nilai_total = float(grading_result.get("nilai_total", 0))
    kesimpulan_singkat = grading_result.get("kesimpulan", "")[:500]

    db.execute(
        text("DELETE FROM scores WHERE submission_id = :submission_id"),
        {"submission_id": submission_id}
    )
    db.execute(
        text("""
            INSERT INTO scores (submission_id, question_number, score, max_score, feedback)
            VALUES (:submission_id, 1, :score, 100, :feedback)
        """),
        {"submission_id": submission_id, "score": nilai_total, "feedback": kesimpulan_singkat}
    )
    db.execute(
        text("UPDATE submissions SET status = 'completed', processed_at = :now WHERE id = :submission_id"),
        {"now": datetime.utcnow(), "submission_id": submission_id}
    )
    db.commit()

    analysis_dir = os.path.join(Config.REPORT_DIR, "analysis")
    os.makedirs(analysis_dir, exist_ok=True)
    analysis_path = os.path.join(analysis_dir, f"submission_{submission_id}.json")
    with open(analysis_path, "w", encoding="utf-8") as f:
        json.dump(grading_result, f, ensure_ascii=False, indent=2)

    return {
        "submission_id": submission_id,
        "student_name": submission_row.student_name,
        "nilai_total": nilai_total,
        "huruf": grading_result.get("huruf"),
        "sections_count": len(grading_result.get("sections", []))
    }


@app.post("/api/grade/submission/{submission_id}/ai")
async def grade_submission_with_ai(submission_id: int, db: Session = Depends(get_db)):
    """Nilai satu submission mahasiswa memakai AI (Gemini vision) berdasarkan
    rubrik + kunci jawaban ujian (format Markdown), lalu simpan skor & buat
    laporan Markdown detail."""
    try:
        ai_service = get_ai_grading_service()
        result = _grade_submission_core(submission_id, db, ai_service)
        return {"status": "success", **result}
    except Exception as e:
        db.rollback()
        logger.error(f"Error in AI grading for submission {submission_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/grade/exam/{exam_id}/all")
async def grade_all_pending_submissions(exam_id: int, db: Session = Depends(get_db)):
    """'Analisis Keseluruhan' — proses semua submission berstatus pending/failed
    pada satu ujian secara berurutan. Kegagalan satu submission tidak
    menghentikan proses submission lain."""
    try:
        ai_service = get_ai_grading_service()
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    pending = db.execute(
        text("SELECT id FROM submissions WHERE exam_id = :exam_id AND status IN ('pending', 'failed')"),
        {"exam_id": exam_id}
    ).fetchall()

    processed = []
    failed = []

    for row in pending:
        try:
            result = _grade_submission_core(row.id, db, ai_service)
            processed.append(result)
        except Exception as e:
            db.rollback()
            db.execute(
                text("UPDATE submissions SET status = 'failed' WHERE id = :id"),
                {"id": row.id}
            )
            db.commit()
            logger.error(f"Gagal menilai submission {row.id} (batch): {e}")
            failed.append({"submission_id": row.id, "error": str(e)})

    return {
        "status": "completed",
        "exam_id": exam_id,
        "total": len(pending),
        "processed": len(processed),
        "failed": len(failed),
        "results": processed,
        "errors": failed
    }


@app.get("/api/report/detailed/exam/{exam_id}/available")
async def list_available_ai_reports(exam_id: int, db: Session = Depends(get_db)):
    """Kembalikan daftar submission_id (dalam satu ujian) yang sudah punya
    hasil analisis AI (bukan sekadar status completed dari mode gratis)."""
    try:
        submission_ids = db.execute(
            text("SELECT id FROM submissions WHERE exam_id = :exam_id"),
            {"exam_id": exam_id}
        ).fetchall()

        available = []
        for row in submission_ids:
            analysis_path = os.path.join(Config.REPORT_DIR, "analysis", f"submission_{row.id}.json")
            if os.path.exists(analysis_path):
                available.append(row.id)

        return {"exam_id": exam_id, "available_submission_ids": available}
    except Exception as e:
        logger.error(f"Error listing available AI reports for exam {exam_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/report/detailed/{submission_id}")
async def get_detailed_report(submission_id: int, db: Session = Depends(get_db)):
    """Download laporan Markdown detail hasil AI grading untuk satu submission."""
    try:
        analysis_path = os.path.join(Config.REPORT_DIR, "analysis", f"submission_{submission_id}.json")

        if not os.path.exists(analysis_path):
            raise HTTPException(
                status_code=404,
                detail="Belum ada analisis AI untuk submission ini. Proses dulu lewat endpoint /api/grade/submission/{id}/ai"
            )

        with open(analysis_path, "r", encoding="utf-8") as f:
            grading_result = json.load(f)

        submission_row = db.execute(
            text("""
                SELECT sub.id, e.title as exam_title,
                       s.name as student_name, s.nim as student_nim,
                       c.name as class_name
                FROM submissions sub
                JOIN exams e ON sub.exam_id = e.id
                JOIN students s ON sub.student_id = s.id
                LEFT JOIN classes c ON e.class_id = c.id
                WHERE sub.id = :submission_id
            """),
            {"submission_id": submission_id}
        ).fetchone()

        if not submission_row:
            raise HTTPException(status_code=404, detail="Submission not found")

        exam_data = {"title": submission_row.exam_title, "class_name": submission_row.class_name or "N/A"}
        student_data = {"name": submission_row.student_name, "nim": submission_row.student_nim}

        report_path = detailed_report_service.generate(exam_data, student_data, grading_result)

        return FileResponse(
            report_path,
            media_type='text/markdown',
            filename=os.path.basename(report_path)
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating detailed report for submission {submission_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _normalize_text(s: str) -> str:
    """Normalisasi teks untuk pencocokan nama (lowercase, hapus spasi berlebih & tanda baca)."""
    s = (s or '').lower().strip()
    s = re.sub(r'[^a-z0-9\s]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s


@app.post("/api/exam/{exam_id}/init-folder")
async def init_exam_folder(exam_id: int, db: Session = Depends(get_db)):
    """Buat folder penyimpanan jawaban untuk suatu ujian (dipanggil sekali,
    tepat setelah ujian dibuat di sisi frontend). Nama folder berbasis judul
    ujian supaya mudah ditelusuri manual, dan disimpan permanen di kolom
    exams.upload_folder supaya tidak berubah walau judul ujian diedit."""
    try:
        exam_row = db.execute(
            text("SELECT id, title, upload_folder FROM exams WHERE id = :exam_id"),
            {"exam_id": exam_id}
        ).fetchone()
        if not exam_row:
            raise HTTPException(status_code=404, detail="Ujian tidak ditemukan.")

        if exam_row.upload_folder:
            # Sudah pernah dibuat sebelumnya, tidak perlu buat ulang
            _get_exam_upload_folder(exam_id, exam_row.title, exam_row.upload_folder)
            return {"status": "success", "folder": exam_row.upload_folder, "already_existed": True}

        folder_name = f"{exam_id}_{_slugify_title(exam_row.title)}"
        _get_exam_upload_folder(exam_id, exam_row.title, folder_name)

        db.execute(
            text("UPDATE exams SET upload_folder = :folder WHERE id = :exam_id"),
            {"folder": folder_name, "exam_id": exam_id}
        )
        db.commit()

        return {"status": "success", "folder": folder_name, "already_existed": False}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating exam folder for exam {exam_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _parse_absen_from_filename(filename: str) -> Optional[int]:
    """Ambil nomor absen dari nama file jawaban mahasiswa. Format yang
    dianjurkan cukup angka absen saja (mis. '12.pdf'), tapi juga tetap
    menerima format dengan tambahan nama di belakangnya (mis.
    '12_Budi_Santoso.pdf' atau '12-Budi Santoso.pdf') — yang penting angka
    absen ada di paling depan nama file."""
    base = os.path.splitext(filename)[0]
    match = re.match(r'^\s*(\d+)', base)
    if not match:
        return None
    return int(match.group(1))


@app.post("/api/upload/bulk-store")
async def bulk_store_by_filename(
    exam_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Upload SATU file jawaban (dipanggil berulang oleh frontend untuk tiap
    file dalam upload massal). TIDAK memanggil AI sama sekali di tahap ini —
    hanya membaca nomor absen dari NAMA FILE (cukup angka absen saja, mis.
    '12.pdf'), mencocokkan ke roster mahasiswa (kolom class_students.absen),
    lalu menyimpan filenya. Penilaian AI baru dilakukan belakangan lewat
    endpoint 'Analisis Keseluruhan' (/api/grade/exam/{exam_id}/all)."""
    try:
        allowed_ext = ('.pdf', '.jpg', '.jpeg', '.png')
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_ext:
            raise HTTPException(status_code=400, detail="Hanya file PDF, JPG, JPEG, atau PNG yang didukung.")

        exam_row = db.execute(
            text("SELECT id, title, class_id, upload_folder FROM exams WHERE id = :exam_id"),
            {"exam_id": exam_id}
        ).fetchone()
        if not exam_row:
            raise HTTPException(status_code=404, detail="Ujian tidak ditemukan.")

        absen = _parse_absen_from_filename(file.filename)
        if absen is None:
            return {
                "filename": file.filename,
                "matched": False,
                "detected_absen": None,
                "message": "Nama file tidak diawali angka absen. Ganti nama file jadi nomor absen saja, mis. '12.pdf'."
            }

        student = db.execute(
            text("""
                SELECT s.id, s.name, s.nim
                FROM students s
                JOIN class_students cs ON cs.student_id = s.id
                WHERE cs.class_id = :class_id AND cs.absen = :absen
            """),
            {"class_id": exam_row.class_id, "absen": absen}
        ).fetchone()

        if not student:
            return {
                "filename": file.filename,
                "matched": False,
                "detected_absen": absen,
                "message": f"Tidak ada mahasiswa dengan absen {absen} di kelas ini. Cek nomor absen atau upload manual lewat form 'Upload Jawaban Mahasiswa'."
            }

        # Simpan file langsung ke folder ujian (dibuat saat ujian dibuat; buat sekarang kalau belum ada)
        exam_folder = _get_exam_upload_folder(exam_id, exam_row.title, exam_row.upload_folder)
        if not exam_row.upload_folder:
            new_folder_name = os.path.basename(exam_folder)
            db.execute(
                text("UPDATE exams SET upload_folder = :folder WHERE id = :exam_id"),
                {"folder": new_folder_name, "exam_id": exam_id}
            )

        final_path = os.path.join(exam_folder, file.filename)
        with open(final_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        existing = db.execute(
            text("SELECT id FROM submissions WHERE exam_id = :exam_id AND student_id = :student_id"),
            {"exam_id": exam_id, "student_id": student.id}
        ).fetchone()

        if existing:
            db.execute(
                text("UPDATE submissions SET answer_sheet_path = :path, status = 'pending', processed_at = NULL WHERE id = :id"),
                {"path": final_path, "id": existing.id}
            )
        else:
            db.execute(
                text("""
                    INSERT INTO submissions (exam_id, student_id, answer_sheet_path, status)
                    VALUES (:exam_id, :student_id, :path, 'pending')
                """),
                {"exam_id": exam_id, "student_id": student.id, "path": final_path}
            )
        db.commit()

        return {
            "filename": file.filename,
            "matched": True,
            "detected_absen": absen,
            "student_id": student.id,
            "student_name": student.name,
            "student_nim": student.nim
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error in bulk store: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/report/exam/{exam_id}/zip")
async def download_all_reports_zip(exam_id: int, db: Session = Depends(get_db)):
    """Download SEMUA laporan Markdown mahasiswa (yang sudah dianalisis AI)
    dalam satu ujian, dikemas jadi satu file ZIP."""
    import zipfile
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    try:
        exam_row = db.execute(
            text("SELECT title FROM exams WHERE id = :exam_id"),
            {"exam_id": exam_id}
        ).fetchone()
        if not exam_row:
            raise HTTPException(status_code=404, detail="Ujian tidak ditemukan.")

        submissions = db.execute(
            text("""
                SELECT sub.id, s.name as student_name, s.nim as student_nim
                FROM submissions sub
                JOIN students s ON sub.student_id = s.id
                WHERE sub.exam_id = :exam_id
                ORDER BY s.name
            """),
            {"exam_id": exam_id}
        ).fetchall()

        zip_buffer = BytesIO()
        included_count = 0

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for sub in submissions:
                analysis_path = os.path.join(Config.REPORT_DIR, "analysis", f"submission_{sub.id}.json")
                if not os.path.exists(analysis_path):
                    continue

                with open(analysis_path, "r", encoding="utf-8") as f:
                    grading_result = json.load(f)

                md_content = detailed_report_service._build_markdown(
                    {"title": exam_row.title, "class_name": ""},
                    {"name": sub.student_name, "nim": sub.student_nim},
                    grading_result
                )

                safe_name = detailed_report_service._slugify(sub.student_name)
                zf.writestr(f"{sub.student_nim}_{safe_name}.md", md_content)
                included_count += 1

        if included_count == 0:
            raise HTTPException(status_code=404, detail="Belum ada mahasiswa yang selesai dianalisis AI di ujian ini.")

        zip_buffer.seek(0)
        safe_title = re.sub(r'[^a-zA-Z0-9_-]', '_', exam_row.title)

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={safe_title}_laporan.zip"}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating ZIP for exam {exam_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/report/exam/{exam_id}/excel")
async def download_scores_excel(exam_id: int, db: Session = Depends(get_db)):
    """Download rekap nilai seluruh mahasiswa dalam satu ujian sebagai file Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    try:
        exam_row = db.execute(
            text("""
                SELECT e.title, c.name as class_name
                FROM exams e
                LEFT JOIN classes c ON e.class_id = c.id
                WHERE e.id = :exam_id
            """),
            {"exam_id": exam_id}
        ).fetchone()
        if not exam_row:
            raise HTTPException(status_code=404, detail="Ujian tidak ditemukan.")

        rows = db.execute(
            text("""
                SELECT s.name as student_name, s.nim as student_nim,
                       sub.status,
                       COALESCE(AVG(sc.score), NULL) as nilai,
                       MAX(sc.feedback) as feedback
                FROM submissions sub
                JOIN students s ON sub.student_id = s.id
                LEFT JOIN scores sc ON sc.submission_id = sub.id
                WHERE sub.exam_id = :exam_id
                GROUP BY s.name, s.nim, sub.status
                ORDER BY s.name
            """),
            {"exam_id": exam_id}
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Rekap Nilai"

        ws['A1'] = exam_row.title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Kelas: {exam_row.class_name or '-'}"
        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')

        headers = ['NIM', 'Nama', 'Status', 'Nilai', 'Catatan Singkat']
        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        status_label = {
            'pending': 'Menunggu', 'processing': 'Diproses',
            'completed': 'Selesai', 'failed': 'Gagal'
        }

        for i, row in enumerate(rows, start=header_row + 1):
            ws.cell(row=i, column=1, value=row.student_nim)
            ws.cell(row=i, column=2, value=row.student_name)
            ws.cell(row=i, column=3, value=status_label.get(row.status, row.status))
            ws.cell(row=i, column=4, value=round(row.nilai, 2) if row.nilai is not None else '-')
            ws.cell(row=i, column=5, value=row.feedback or '-')

        for col_letter, width in [('A', 18), ('B', 30), ('C', 14), ('D', 10), ('E', 50)]:
            ws.column_dimensions[col_letter].width = width

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        safe_title = re.sub(r'[^a-zA-Z0-9_-]', '_', exam_row.title)

        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={safe_title}_nilai.xlsx"}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating Excel for exam {exam_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))
